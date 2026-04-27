# -*- coding: utf-8 -*-
"""Pipeline del catálogo maestro de estaciones DHIME."""

from __future__ import annotations

import logging
import zipfile
from dataclasses import dataclass
from datetime import date
from difflib import SequenceMatcher
from pathlib import Path
from typing import Iterable

from ideam_dhime.station_catalog import StationMetadata, normalize_location_name

logger = logging.getLogger("ideam_dhime")

CANONICAL_COLUMNS = [
    "station_code",
    "station_name",
    "depto_original",
    "municipio_original",
    "depto_oficial",
    "municipio_oficial",
    "entidad",
    "fecha_ini_op",
    "fecha_fin_op",
    "latitud",
    "longitud",
    "categoria_estacion",
    "metodo_match_depto",
    "score_depto",
    "metodo_match_municipio",
    "score_municipio",
]

SOURCE_COLUMNS_KEY = "__source_columns__"
SOURCE_ROW_KEY = "__source_row__"

RESOURCE_PRIMARY_EXCEL = "CATALOGO_NACIONAL_DE_ESTACIONES_(EXCEL).xls"
RESOURCE_OTHER_EXCEL = "CATALOGO_NACIONAL_DE_OTRAS_ENTIDADES_(EXCEL).xls"
RESOURCE_SHAPE_ZIP = "CATALOGO_NACIONAL_DE_ESTACIONES_(SHAPE).zip"
RESOURCE_SHAPE_DBF = "CATALOGO_NACIONAL_DE_ESTACIONES_(SHAPE).dbf"

REQUIRED_STATION_RESOURCES = (
    RESOURCE_PRIMARY_EXCEL,
    RESOURCE_OTHER_EXCEL,
    RESOURCE_SHAPE_ZIP,
)

RESOURCE_DOWNLOAD_ALIASES = {
    RESOURCE_PRIMARY_EXCEL: (
        "CATALOGO_NACIONAL_DE_ESTACIONES_(EXCEL).xls",
        "CNE.xls",
        "CNE_IDEAM.xls",
        "CNE_E.xls",
        "CNE_IDEAM_EXCEL.xls",
    ),
    RESOURCE_OTHER_EXCEL: (
        "CATALOGO_NACIONAL_DE_OTRAS_ENTIDADES_(EXCEL).xls",
        "CNE_OE.xls",
        "CNE_OTRAS_ENTIDADES.xls",
        "CNE_OTRAS_ENTIDADES_EXCEL.xls",
    ),
    RESOURCE_SHAPE_ZIP: (
        "CATALOGO_NACIONAL_DE_ESTACIONES_(SHAPE).zip",
        "CNE_SHAPE.zip",
        "CNE_SHP.zip",
        "CNE.zip",
    ),
    RESOURCE_SHAPE_DBF: (
        "CATALOGO_NACIONAL_DE_ESTACIONES_(SHAPE).dbf",
        "CNE_SHAPE.dbf",
        "CNE_SHP.dbf",
        "CNE.dbf",
    ),
}

_ALIASES = {
    "station_code": ("CODIGO", "CODIGO ESTACION", "CODIGO_ESTACION", "station_code"),
    "station_name": ("NOMBRE", "NOMBRE ESTACION", "NOMBRE_ESTACION", "station_name"),
    "depto": ("DEPARTAMENTO", "DEPTO", "DEPARTAMEN", "depto"),
    "municipio": ("MUNICIPIO", "MUNICIP", "municipio"),
    "fecha_ini_op": ("FECHA INICIO", "FECHA_INICIO", "FECHA INI OP", "fecha_ini_op"),
    "fecha_fin_op": (
        "FECHA FIN",
        "FECHA_FIN",
        "FECHA FIN OP",
        "FECHA_SUSPENSION",
        "fecha_fin_op",
    ),
    "latitud": ("LATITUD", "LAT", "latitud"),
    "longitud": ("LONGITUD", "LON", "LONG", "longitud"),
    "categoria_estacion": ("CATEGORIA", "CATEGORIA ESTACION", "categoria_estacion"),
    "entidad": ("ENTIDAD", "entidad"),
}


@dataclass(frozen=True)
class CatalogReport:
    """Resumen de una generación de catálogo."""

    rows_total: int
    quality_rows: int
    output_csv: Path
    stations_py: Path | None = None
    locations_py: Path | None = None
    quality_csv: Path | None = None


def _norm_key(value: str) -> str:
    return normalize_location_name(value).replace("_", " ")


def _first_value(row: dict[str, object], aliases: Iterable[str]) -> str:
    normalized = {_norm_key(str(key)): value for key, value in row.items()}
    for alias in aliases:
        value = normalized.get(_norm_key(alias))
        if value is None:
            continue
        return str(value).strip()
    return ""


def _canonical_station_row(row: dict[str, object], default_entity: str) -> dict[str, str]:
    entity = _first_value(row, _ALIASES["entidad"]) or default_entity
    return {
        SOURCE_COLUMNS_KEY: list(row.keys()),  # type: ignore[dict-item]
        SOURCE_ROW_KEY: {str(key): value for key, value in row.items()},  # type: ignore[dict-item]
        "station_code": _first_value(row, _ALIASES["station_code"]),
        "station_name": _first_value(row, _ALIASES["station_name"]),
        "depto_original": _first_value(row, _ALIASES["depto"]),
        "municipio_original": _first_value(row, _ALIASES["municipio"]),
        "depto_oficial": "",
        "municipio_oficial": "",
        "entidad": entity,
        "fecha_ini_op": _first_value(row, _ALIASES["fecha_ini_op"]),
        "fecha_fin_op": _first_value(row, _ALIASES["fecha_fin_op"]),
        "latitud": _first_value(row, _ALIASES["latitud"]),
        "longitud": _first_value(row, _ALIASES["longitud"]),
        "categoria_estacion": _first_value(row, _ALIASES["categoria_estacion"]),
        "tecnologia": _first_value(row, ("TECNOLOGIA", "tecnologia")),
        "metodo_match_depto": "",
        "score_depto": "",
        "metodo_match_municipio": "",
        "score_municipio": "",
    }


def compile_station_rows(
    primary_rows: Iterable[dict[str, object]],
    other_rows: Iterable[dict[str, object]],
) -> list[dict[str, str]]:
    """Une los dos CNE y deduplica por código con precedencia del CNE IDEAM."""
    compiled: dict[str, dict[str, str]] = {}
    for raw in other_rows:
        row = _canonical_station_row(raw, default_entity="OTRAS_ENTIDADES")
        code = row["station_code"].strip()
        if code:
            compiled[code] = row

    for raw in primary_rows:
        row = _canonical_station_row(raw, default_entity="IDEAM")
        code = row["station_code"].strip()
        if code:
            compiled[code] = row

    return [compiled[code] for code in sorted(compiled)]


def read_excel_rows(excel_path: str | Path) -> list[dict[str, object]]:
    path = Path(excel_path)
    if path.suffix.lower() == ".xlsx":
        try:
            from openpyxl import load_workbook
        except ImportError as exc:
            raise ImportError("Para leer Excel .xlsx instala 'openpyxl'.") from exc

        workbook = load_workbook(path, read_only=True, data_only=True)
        sheet = workbook.worksheets[0]
        iterator = sheet.iter_rows(values_only=True)
        header_values = next(iterator, None)
        if header_values is None:
            return []
        header = [str(value or "").strip() for value in header_values]
        return [
            {column: value for column, value in zip(header, row_values) if column}
            for row_values in iterator
        ]

    try:
        import xlrd
    except ImportError as exc:
        raise ImportError("Para leer Excel .xls instala 'xlrd'.") from exc

    workbook = xlrd.open_workbook(path)
    sheet = workbook.sheet_by_index(0)
    if sheet.nrows == 0:
        return []

    header = [str(sheet.cell_value(0, col)).strip() for col in range(sheet.ncols)]
    rows: list[dict[str, object]] = []
    for row_idx in range(1, sheet.nrows):
        row: dict[str, object] = {}
        for col_idx, column in enumerate(header):
            if not column:
                continue
            value = sheet.cell_value(row_idx, col_idx)
            if sheet.cell_type(row_idx, col_idx) == xlrd.XL_CELL_NUMBER:
                value = str(int(value)) if float(value).is_integer() else str(value)
            row[column] = value
        rows.append(row)
    return rows


def _match_official(value: str, choices: Iterable[str], cutoff: float = 0.92) -> tuple[str, str, str]:
    choices_list = sorted(set(choices))
    if not value:
        return "", "unmatched", "0.000"
    if not choices_list:
        return value, "unmatched", "0.000"

    if value in choices_list:
        return value, "exact", "1.000"

    target = normalize_location_name(value)
    normalized_map = {normalize_location_name(choice): choice for choice in choices_list}
    if target in normalized_map:
        return normalized_map[target], "normalized", "1.000"

    parenthetical_map = {
        normalize_location_name(choice.split("(", 1)[0].strip()): choice for choice in choices_list
    }
    if target in parenthetical_map:
        return parenthetical_map[target], "normalized_parenthetical", "1.000"

    best_choice = ""
    best_score = 0.0
    for normalized_choice, original_choice in normalized_map.items():
        score = SequenceMatcher(None, target, normalized_choice).ratio()
        if score > best_score:
            best_choice = original_choice
            best_score = score
    if best_score >= cutoff:
        return best_choice, "fuzzy", f"{best_score:.3f}"
    return value, "unmatched", f"{best_score:.3f}"


def _is_automatic_telemetry_station(row: dict[str, str]) -> bool:
    source_row = row.get(SOURCE_ROW_KEY, {})  # type: ignore[assignment]
    technology = ""
    if isinstance(source_row, dict):
        technology = _first_value(source_row, ("TECNOLOGIA", "tecnologia"))
    technology = technology or row.get("tecnologia", "")
    normalized = normalize_location_name(technology)
    return "AUTOMATICA CON TELEMETRIA" in normalized or "AUTOMATICA SIN TELEMETRIA" in normalized


def _parse_date_ddmmyyyy(value: str) -> date | None:
    raw = (value or "").strip()
    if not raw:
        return None
    for separator in ("/", "-"):
        if separator not in raw:
            continue
        parts = raw.split(separator)
        if len(parts) != 3:
            continue
        try:
            if len(parts[0]) == 4:
                year, month, day = (int(part) for part in parts)
            else:
                day, month, year = (int(part) for part in parts)
            return date(year, month, day)
        except ValueError:
            continue
    return None


def _is_suspended_before_1970(row: dict[str, str]) -> bool:
    source_row = row.get(SOURCE_ROW_KEY, {})  # type: ignore[assignment]
    value = ""
    if isinstance(source_row, dict):
        value = _first_value(source_row, _ALIASES["fecha_fin_op"])
    value = value or row.get("fecha_fin_op", "")
    parsed = _parse_date_ddmmyyyy(value)
    return parsed is not None and parsed < date(1970, 1, 1)


def correct_station_locations(
    rows: Iterable[dict[str, str]],
    official_locations: Iterable[tuple[str, str]],
    station_locations: dict[str, dict[str, str]] | None = None,
    cutoff: float = 0.92,
    trust_cne_for_automatic_telemetry: bool = True,
    trust_cne_for_pre_1970_suspended: bool = True,
) -> tuple[list[dict[str, str]], list[dict[str, str]]]:
    """Corrige departamento y luego municipio filtrando por departamento corregido."""
    locations = list(official_locations)
    station_lookup = station_locations or {}
    departments = {department for department, _municipality in locations}
    municipalities_by_department: dict[str, set[str]] = {}
    for department, municipality in locations:
        municipalities_by_department.setdefault(department, set()).add(municipality)

    corrected: list[dict[str, str]] = []
    quality: list[dict[str, str]] = []
    for row in rows:
        out = dict(row)
        station_code = row.get("station_code", "").strip()
        station_official = station_lookup.get(station_code)
        if trust_cne_for_automatic_telemetry and _is_automatic_telemetry_station(row):
            depto = row.get("depto_original", "")
            municipio = row.get("municipio_original", "")
            depto_method = "cne_automatic_telemetry"
            municipio_method = "cne_automatic_telemetry"
            depto_score = "1.000"
            municipio_score = "1.000"
        elif trust_cne_for_pre_1970_suspended and _is_suspended_before_1970(row):
            depto = row.get("depto_original", "")
            municipio = row.get("municipio_original", "")
            depto_method = "cne_suspended_pre_1970"
            municipio_method = "cne_suspended_pre_1970"
            depto_score = "1.000"
            municipio_score = "1.000"
        elif station_official:
            depto = station_official.get("department", "")
            municipio = station_official.get("municipality", "")
            depto_method = "station_code"
            municipio_method = "station_code"
            depto_score = "1.000"
            municipio_score = "1.000"
            out.update(
                {
                    "station_name": station_official.get("station_name") or out.get("station_name", ""),
                    "latitud": station_official.get("latitud") or out.get("latitud", ""),
                    "longitud": station_official.get("longitud") or out.get("longitud", ""),
                    "categoria_estacion": station_official.get("categoria_estacion")
                    or out.get("categoria_estacion", ""),
                }
            )
        else:
            depto, depto_method, depto_score = _match_official(
                row.get("depto_original", ""),
                departments,
                cutoff=cutoff,
            )
            municipality_choices = municipalities_by_department.get(depto, set())
            municipio, municipio_method, municipio_score = _match_official(
                row.get("municipio_original", ""),
                municipality_choices,
                cutoff=cutoff,
            )
        out.update(
            {
                "depto_oficial": depto,
                "municipio_oficial": municipio,
                "metodo_match_depto": depto_method,
                "score_depto": depto_score,
                "metodo_match_municipio": municipio_method,
                "score_municipio": municipio_score,
            }
        )
        corrected.append(out)
        if depto_method == "unmatched" or municipio_method == "unmatched":
            quality.append(
                {
                    "station_code": out.get("station_code", ""),
                    "depto_original": out.get("depto_original", ""),
                    "municipio_original": out.get("municipio_original", ""),
                    "depto_oficial": depto,
                    "municipio_oficial": municipio,
                    "metodo_match_depto": depto_method,
                    "score_depto": depto_score,
                    "metodo_match_municipio": municipio_method,
                    "score_municipio": municipio_score,
                }
            )
    return corrected, quality


def extract_dbf_from_shape_zip(shape_zip: str | Path, output_dir: str | Path) -> Path:
    """Descomprime el ZIP de shape, elimina el ZIP y conserva solo el DBF."""
    zip_path = Path(shape_zip)
    out = Path(output_dir)
    out.mkdir(parents=True, exist_ok=True)

    with zipfile.ZipFile(zip_path, "r") as zf:
        zf.extractall(out)

    dbf_files = list(out.glob("*.dbf")) + list(out.glob("*.DBF"))
    if not dbf_files:
        raise FileNotFoundError(f"No se encontró DBF dentro de {zip_path}")
    dbf_path = dbf_files[0]

    for path in out.iterdir():
        if path == dbf_path:
            continue
        if path.is_file():
            path.unlink()
    zip_path.unlink()
    return dbf_path


def _classify_downloaded_resource(path: Path) -> str | None:
    filename = normalize_location_name(path.name)
    for resource_name, aliases in RESOURCE_DOWNLOAD_ALIASES.items():
        if any(filename == normalize_location_name(alias) for alias in aliases):
            return resource_name

    suffix = path.suffix.lower()
    stem = normalize_location_name(path.stem)
    if suffix == ".zip" and ("SHAPE" in stem or "SHP" in stem or stem == "CNE"):
        return RESOURCE_SHAPE_ZIP
    if suffix == ".dbf" and ("SHAPE" in stem or "SHP" in stem or stem == "CNE"):
        return RESOURCE_SHAPE_DBF
    if suffix == ".xls" and ("OE" in stem or "OTRAS" in stem):
        return RESOURCE_OTHER_EXCEL
    if suffix == ".xls" and ("CNE" in stem or "ESTACIONES" in stem):
        return RESOURCE_PRIMARY_EXCEL
    return None


def find_local_station_catalog_resources(source_dir: str | Path) -> dict[str, Path]:
    """
    Ubica en `source_dir` los dos Excel CNE y el DBF/ZIP del shape.

    Convención recomendada: crear una carpeta `CNE/` con:
    - `CATALOGO_NACIONAL_DE_ESTACIONES_(EXCEL).xls`
    - `CATALOGO_NACIONAL_DE_OTRAS_ENTIDADES_(EXCEL).xls`
    - `CNE.dbf` o `CATALOGO_NACIONAL_DE_ESTACIONES_(SHAPE).dbf`
    """
    base = Path(source_dir).absolute()
    if not base.exists():
        raise FileNotFoundError(
            f"No existe la carpeta CNE esperada: {base}. "
            "Crea una carpeta CNE con los dos Excel y el DBF del shape."
        )

    found: dict[str, Path] = {}
    excel_candidates: list[Path] = []
    dbf_candidates: list[Path] = []
    for path in base.rglob("*"):
        if not path.is_file():
            continue
        if path.suffix.lower() in (".xls", ".xlsx"):
            excel_candidates.append(path)
        if path.suffix.lower() == ".dbf":
            dbf_candidates.append(path)
        resource_name = _classify_downloaded_resource(path)
        if resource_name is not None:
            found[resource_name] = path

    if (
        (RESOURCE_PRIMARY_EXCEL not in found or RESOURCE_OTHER_EXCEL not in found)
        and len(excel_candidates) == 2
    ):
        sorted_excels = sorted(excel_candidates, key=lambda path: path.name.lower())
        other = next(
            (
                path
                for path in sorted_excels
                if any(token in normalize_location_name(path.name) for token in ("OE", "OTRAS"))
            ),
            None,
        )
        if other is not None:
            primary = next(path for path in sorted_excels if path != other)
        else:
            primary, other = sorted_excels
        found.setdefault(RESOURCE_PRIMARY_EXCEL, primary)
        found.setdefault(RESOURCE_OTHER_EXCEL, other)

    if RESOURCE_SHAPE_DBF not in found and len(dbf_candidates) == 1:
        found[RESOURCE_SHAPE_DBF] = dbf_candidates[0]

    required = (RESOURCE_PRIMARY_EXCEL, RESOURCE_OTHER_EXCEL)
    missing = [name for name in required if name not in found]
    if RESOURCE_SHAPE_DBF not in found and RESOURCE_SHAPE_ZIP not in found:
        missing.append(RESOURCE_SHAPE_DBF)
    if missing:
        present = sorted(str(path.relative_to(base)) for path in base.rglob("*") if path.is_file())
        raise FileNotFoundError(
            "Faltan archivos requeridos en la carpeta CNE. "
            f"Faltantes={missing}; presentes={present}"
        )
    return found


def read_official_locations_from_dbf(dbf_path: str | Path) -> list[tuple[str, str]]:
    """Lee ubicaciones oficiales desde DBF usando dbfread si está instalado."""
    try:
        from dbfread import DBF
    except ImportError as exc:
        raise ImportError(
            "Para leer DBF instala la dependencia opcional 'dbfread'."
        ) from exc

    encoding = _detect_dbf_encoding(Path(dbf_path), DBF)
    rows = DBF(str(dbf_path), load=True, encoding=encoding)
    locations: set[tuple[str, str]] = set()
    for row in rows:
        department = _first_value(
            row,
            ("d_DEPARTAM", "DEPARTAMENTO", "DEPTO", "DPTO_CNMBR", "NOM_DEP"),
        )
        municipality = _first_value(
            row,
            ("d_MUNICIPI", "MUNICIPIO", "MPIO_CNMBR", "NOM_MUN"),
        )
        if department and municipality:
            locations.add((department, municipality))
    return sorted(locations)


def read_official_station_metadata_from_dbf(dbf_path: str | Path) -> dict[str, dict[str, str]]:
    """Lee metadatos oficiales por código de estación desde el DBF."""
    try:
        from dbfread import DBF
    except ImportError as exc:
        raise ImportError(
            "Para leer DBF instala la dependencia opcional 'dbfread'."
        ) from exc

    encoding = _detect_dbf_encoding(Path(dbf_path), DBF)
    rows = DBF(str(dbf_path), load=True, encoding=encoding)
    metadata: dict[str, dict[str, str]] = {}
    for row in rows:
        code = _first_value(row, ("CODIGO", "station_code"))
        department = _first_value(
            row,
            ("d_DEPARTAM", "DEPARTAMENTO", "DEPTO", "DPTO_CNMBR", "NOM_DEP"),
        )
        municipality = _first_value(
            row,
            ("d_MUNICIPI", "MUNICIPIO", "MPIO_CNMBR", "NOM_MUN"),
        )
        if not code or not department or not municipality:
            continue
        metadata[code] = {
            "station_name": _first_value(row, ("nombre", "NOMBRE", "station_name")),
            "department": department,
            "municipality": municipality,
            "latitud": _first_value(row, ("latitud", "LATITUD", "LAT")),
            "longitud": _first_value(row, ("longitud", "LONGITUD", "LON", "LONG")),
            "categoria_estacion": _first_value(row, ("d_CATEGORI", "CATEGORIA")),
        }
    return metadata


def _read_cpg_encoding(dbf_path: Path) -> str | None:
    cpg_path = dbf_path.with_suffix(".cpg")
    if not cpg_path.exists():
        return None
    value = cpg_path.read_text(encoding="ascii", errors="ignore").strip()
    aliases = {
        "1252": "cp1252",
        "ANSI 1252": "cp1252",
        "UTF-8": "utf-8",
        "LATIN1": "latin-1",
    }
    return aliases.get(value.upper(), value or None)


def _text_quality_score(values: Iterable[str]) -> int:
    joined = " ".join(values)
    bad_markers = ("Ã", "Â", "�", "Ð", "¥", "├", "┬", "┴", "┼", "│", "┤")
    return sum(joined.count(marker) for marker in bad_markers)


def _detect_dbf_encoding(dbf_path: Path, dbf_cls) -> str:
    candidates = []
    cpg_encoding = _read_cpg_encoding(dbf_path)
    if cpg_encoding:
        candidates.append(cpg_encoding)
    candidates.extend(["cp1252", "latin-1", "utf-8", "cp850"])

    best_encoding = "cp1252"
    best_score: tuple[int, int] | None = None
    for encoding in dict.fromkeys(candidates):
        try:
            table = dbf_cls(str(dbf_path), load=False, encoding=encoding)
            values: list[str] = []
            for idx, row in enumerate(table):
                values.extend(str(value) for value in row.values() if isinstance(value, str))
                if idx >= 25:
                    break
            score = (_text_quality_score(values), -len(" ".join(values)))
        except Exception:
            continue
        if best_score is None or score < best_score:
            best_encoding = encoding
            best_score = score
    return best_encoding


def _write_rows_xlsx(
    rows: Iterable[dict[str, object]],
    output_path: str | Path,
    columns: list[str],
    sheet_name: str,
) -> Path:
    try:
        from openpyxl import Workbook
    except ImportError as exc:
        raise ImportError("Para escribir archivos Excel instala 'openpyxl'.") from exc

    path = Path(output_path)
    path.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = sheet_name
    sheet.append(columns)
    for row in rows:
        sheet.append([row.get(column, "") for column in columns])
    sheet.freeze_panes = "A2"
    workbook.save(path)
    return path


def _source_output_row(row: dict[str, str]) -> dict[str, object]:
    source_row = dict(row.get(SOURCE_ROW_KEY, {}))  # type: ignore[arg-type]
    if not source_row:
        return {column: row.get(column, "") for column in CANONICAL_COLUMNS}

    department_column = next(
        (column for column in source_row if _norm_key(column) in {_norm_key(alias) for alias in _ALIASES["depto"]}),
        None,
    )
    municipality_column = next(
        (
            column
            for column in source_row
            if _norm_key(column) in {_norm_key(alias) for alias in _ALIASES["municipio"]}
        ),
        None,
    )
    if department_column:
        source_row[department_column] = row.get("depto_oficial") or row.get("depto_original", "")
    if municipality_column:
        source_row[municipality_column] = row.get("municipio_oficial") or row.get(
            "municipio_original", ""
        )
    return source_row


def _source_output_columns(rows: list[dict[str, str]]) -> list[str]:
    for row in rows:
        columns = row.get(SOURCE_COLUMNS_KEY)
        if isinstance(columns, list) and columns:
            return [str(column) for column in columns]
    return CANONICAL_COLUMNS


def write_station_catalog_excel(rows: Iterable[dict[str, str]], output_xlsx: str | Path) -> Path:
    """Escribe el Excel final conservando la estructura original del CNE."""
    row_list = list(rows)
    columns = _source_output_columns(row_list)
    output_rows = [_source_output_row(row) for row in row_list]
    return _write_rows_xlsx(output_rows, output_xlsx, columns, "guia_estaciones")


def write_quality_excel(rows: Iterable[dict[str, str]], output_xlsx: str | Path) -> Path:
    """Escribe reporte Excel de coincidencias no resueltas o dudosas."""
    columns = [
        "station_code",
        "depto_original",
        "municipio_original",
        "depto_oficial",
        "municipio_oficial",
        "metodo_match_depto",
        "score_depto",
        "metodo_match_municipio",
        "score_municipio",
    ]
    return _write_rows_xlsx(list(rows), output_xlsx, columns, "calidad_catalogo")


def write_station_snapshots(
    rows: Iterable[dict[str, str]],
    official_locations: Iterable[tuple[str, str]],
    stations_py: str | Path,
    locations_py: str | Path,
    *,
    generated_at: str | None = None,
    sources: dict[str, str] | None = None,
) -> tuple[Path, Path]:
    """Genera snapshots Python versionables para runtime."""
    generated = generated_at or date.today().isoformat()
    source_map = sources or {}
    stations_path = Path(stations_py)
    locations_path = Path(locations_py)
    stations_path.parent.mkdir(parents=True, exist_ok=True)
    locations_path.parent.mkdir(parents=True, exist_ok=True)

    station_lines = [
        "# -*- coding: utf-8 -*-",
        '"""Snapshot generado del catálogo maestro de estaciones DHIME."""',
        "",
        "from __future__ import annotations",
        "",
        "from typing import Final",
        "",
        "from ideam_dhime.station_catalog import StationMetadata",
        "",
        f"CATALOG_GENERATED_AT: Final[str] = {generated!r}",
        f"CATALOG_SOURCES: Final[dict[str, str]] = {source_map!r}",
        "",
        "STATIONS_DHIME: Final[dict[str, StationMetadata]] = {",
    ]
    for row in sorted(rows, key=lambda item: item.get("station_code", "")):
        code = row.get("station_code", "")
        if not code:
            continue
        metadata = StationMetadata(
            station_code=code,
            station_name=row.get("station_name", ""),
            department=row.get("depto_oficial", ""),
            municipality=row.get("municipio_oficial", ""),
            entity=row.get("entidad", ""),
            fecha_ini_op=row.get("fecha_ini_op", ""),
            fecha_fin_op=row.get("fecha_fin_op", ""),
            latitude=row.get("latitud", ""),
            longitude=row.get("longitud", ""),
            category=row.get("categoria_estacion", ""),
        )
        station_lines.append(f"    {code!r}: {metadata!r},")
    station_lines.append("}")
    station_lines.append("")
    stations_path.write_text("\n".join(station_lines), encoding="utf-8")

    location_lines = [
        "# -*- coding: utf-8 -*-",
        '"""Snapshot generado de ubicaciones oficiales derivadas del DBF."""',
        "",
        "from __future__ import annotations",
        "",
        "from typing import Final",
        "",
        f"CATALOG_GENERATED_AT: Final[str] = {generated!r}",
        f"CATALOG_SOURCES: Final[dict[str, str]] = {source_map!r}",
        "",
        "OFFICIAL_LOCATIONS: Final[tuple[tuple[str, str], ...]] = (",
    ]
    for department, municipality in sorted(set(official_locations)):
        location_lines.append(f"    ({department!r}, {municipality!r}),")
    location_lines.append(")")
    location_lines.append("")
    locations_path.write_text("\n".join(location_lines), encoding="utf-8")
    return stations_path, locations_path


def build_station_catalog_from_rows(
    primary_rows: Iterable[dict[str, object]],
    other_rows: Iterable[dict[str, object]],
    official_locations: Iterable[tuple[str, str]],
    output_catalog: str | Path,
    *,
    station_locations: dict[str, dict[str, str]] | None = None,
    trust_cne_for_automatic_telemetry: bool = True,
    trust_cne_for_pre_1970_suspended: bool = True,
    quality_report: str | Path | None = None,
    stations_py: str | Path | None = None,
    locations_py: str | Path | None = None,
    generated_at: str | None = None,
    sources: dict[str, str] | None = None,
) -> CatalogReport:
    """Construye el catálogo final desde filas ya leídas de los CNE y DBF."""
    compiled = compile_station_rows(primary_rows, other_rows)
    corrected, quality_rows = correct_station_locations(
        compiled,
        official_locations,
        station_locations=station_locations,
        trust_cne_for_automatic_telemetry=trust_cne_for_automatic_telemetry,
        trust_cne_for_pre_1970_suspended=trust_cne_for_pre_1970_suspended,
    )
    output_path = write_station_catalog_excel(corrected, output_catalog)
    quality_path = write_quality_excel(quality_rows, quality_report) if quality_report else None
    stations_path = None
    locations_path = None
    if stations_py and locations_py:
        stations_path, locations_path = write_station_snapshots(
            corrected,
            official_locations,
            stations_py,
            locations_py,
            generated_at=generated_at,
            sources=sources,
        )
    return CatalogReport(
        rows_total=len(corrected),
        quality_rows=len(quality_rows),
        output_csv=output_path,
        stations_py=stations_path,
        locations_py=locations_path,
        quality_csv=quality_path,
    )


def build_station_catalog_from_files(
    primary_excel: str | Path,
    other_excel: str | Path,
    shape_source: str | Path,
    output_catalog: str | Path,
    *,
    quality_report: str | Path | None = None,
    stations_py: str | Path | None = None,
    locations_py: str | Path | None = None,
    generated_at: str | None = None,
    sources: dict[str, str] | None = None,
    cleanup_sources: bool = False,
    trust_cne_for_automatic_telemetry: bool = True,
    trust_cne_for_pre_1970_suspended: bool = True,
) -> CatalogReport:
    """Construye el catálogo final desde los dos Excel oficiales y un DBF/ZIP del shape."""
    primary_path = Path(primary_excel)
    other_path = Path(other_excel)
    shape_path = Path(shape_source)
    primary_rows = read_excel_rows(primary_path)
    other_rows = read_excel_rows(other_path)
    if shape_path.suffix.lower() == ".zip":
        dbf_dir = shape_path.with_suffix("")
        dbf_path = extract_dbf_from_shape_zip(shape_path, dbf_dir)
    else:
        dbf_path = shape_path
    locations = read_official_locations_from_dbf(dbf_path)
    station_locations = read_official_station_metadata_from_dbf(dbf_path)
    report = build_station_catalog_from_rows(
        primary_rows,
        other_rows,
        locations,
        output_catalog,
        station_locations=station_locations,
        trust_cne_for_automatic_telemetry=trust_cne_for_automatic_telemetry,
        trust_cne_for_pre_1970_suspended=trust_cne_for_pre_1970_suspended,
        quality_report=quality_report,
        stations_py=stations_py,
        locations_py=locations_py,
        generated_at=generated_at,
        sources=sources
        or {
            "primary_excel": primary_path.name,
            "other_excel": other_path.name,
            "shape_dbf": dbf_path.name,
        },
    )
    if cleanup_sources:
        for path in (primary_path, other_path, shape_path):
            try:
                path.unlink()
            except FileNotFoundError:
                pass
    return report


def regenerate_station_catalog(
    source_dir: str | Path | None = None,
    *,
    output_root: str | Path | None = None,
    cleanup_sources: bool = False,
    trust_cne_for_automatic_telemetry: bool = True,
    trust_cne_for_pre_1970_suspended: bool = True,
) -> CatalogReport:
    """
    Regenera el catálogo maestro completo desde una carpeta local `CNE/`.

    El usuario avanzado/mantenedor descarga manualmente desde Recursos DHIME los
    dos Excel CNE y el DBF del shape, los guarda en `CNE/`, y ejecuta este comando.
    """
    root = Path(output_root).absolute() if output_root else Path(__file__).resolve().parents[1]
    source = Path(source_dir).absolute() if source_dir else Path.cwd() / "CNE"
    resources = find_local_station_catalog_resources(source)
    shape_source = resources.get(RESOURCE_SHAPE_DBF) or resources[RESOURCE_SHAPE_ZIP]
    generated_at = date.today().isoformat()
    docs_dir = root / "docs"
    package_dir = root / "ideam_dhime"

    return build_station_catalog_from_files(
        resources[RESOURCE_PRIMARY_EXCEL],
        resources[RESOURCE_OTHER_EXCEL],
        shape_source,
        docs_dir / "guia_estaciones.xlsx",
        quality_report=docs_dir / "guia_estaciones_calidad.xlsx",
        stations_py=package_dir / "stations_generated.py",
        locations_py=package_dir / "locations_generated.py",
        generated_at=generated_at,
        sources={name: path.name for name, path in resources.items()},
        cleanup_sources=cleanup_sources,
        trust_cne_for_automatic_telemetry=trust_cne_for_automatic_telemetry,
        trust_cne_for_pre_1970_suspended=trust_cne_for_pre_1970_suspended,
    )
